import os, re, tempfile, subprocess, sys
import pandas as pd
import pdfplumber
from dateutil import parser as dateparser

# 表格抽取
try:
    import camelot
except Exception:
    camelot = None
try:
    import tabula
except Exception:
    tabula = None

# OCR 相关
try:
    import pytesseract
    from pdf2image import convert_from_path
    import cv2
    import numpy as np
except Exception:
    pytesseract = None
try:
    import img2pdf
except Exception:
    img2pdf = None

WECHAT_HEADERS = [
    "交易单号","交易时间","交易类型","收/支/其他","交易方式","金额(元)","交易对方","商户单号"
]
# 支付宝常见列名（不同版式会有差异，列名取子串匹配）
ALIPAY_HEADERS = [
    "交易时间","业务类型","收/支","对方账号","对方账户","对方名称","交易对手","商品说明",
    "金额","资金渠道","收/付款方式","交易号","订单号","商户订单号","备注","余额","手续费"
]

# 泛化：各类银行/平台流水常见列名线索（用于识别与映射）
DATE_HEADERS = [
    "交易时间","交易日期","记账日期","入账日期","发生日期","业务日期","日期","时间"
]
AMOUNT_HEADERS = [
    "金额(元)","金额","发生额","交易金额","交易发生额","入账金额","记账金额","借方金额","贷方金额",
    "收入金额","支出金额","借方发生额","贷方发生额","借(入)","贷(出)"
]
BALANCE_HEADERS = [
    "余额","账户余额","当前余额","可用余额"
]
DESCRIPTION_HEADERS = [
    "摘要","用途","附言","备注","说明","商品说明","交易摘要"
]
COUNTERPARTY_HEADERS = [
    "交易对方","对方账户","对方账号","对方名称","交易对手","收款人","付款人","对方户名","对手账户","交易对象"
]
DIRECTION_HEADERS = [
    "收/支/其他","收支","收支方向","借贷方向","借/贷","方向","收/支"
]
METHOD_HEADERS = [
    "交易方式","收/付款方式","交易渠道","资金渠道","渠道","支付方式","交易类型","业务类型"
]
TXN_ID_HEADERS = [
    "交易单号","交易号","流水号","参考号","交易参考号","交易流水号","凭证号"
]
MERCHANT_ID_HEADERS = [
    "商户单号","商户订单号","订单号","订单编号"
]

# 银行对账单常见表头（泛化，包含建行等）
BANK_GENERIC_HEADERS = [
    "序号","币种","钞/汇","钞汇","记账日期","交易日期","摘要","借/贷","借贷","交易金额","金额",
    "账户余额","余额","对手账号","对手户名","对手账号/户名","交易地点/渠道","地点/渠道","渠道","网点",
]

# ------------------ 公共清洗 ------------------

def _normalize_cols(cols):
    """Lightweight normalization for detection/matching only.
    Keeps a strict version for heuristics but should NOT be used for export.
    - Trim ends
    - Remove internal spaces and newlines for robust matching
    """
    return [str(c).strip().replace("\n", "").replace(" ", "") for c in cols]

def _raw_header_cells(cols):
    """Preserve original header text for export.
    Only strip leading/trailing spaces; keep inner spaces/newlines as-is.
    """
    return [None if pd.isna(c) else str(c).strip() for c in cols]

def _is_wechat_table(df):
    cols = _normalize_cols(df.columns)
    hit = sum(1 for h in WECHAT_HEADERS if any(h in c for c in cols))
    return hit >= 5

def _is_alipay_table(df):
    cols = _normalize_cols(df.columns)
    hit = sum(1 for h in ALIPAY_HEADERS if any(h in c for c in cols))
    # 支付宝表一般也有金额/交易时间/收支/交易号等，阈值 5 足够稳妥
    return hit >= 5

def _looks_like_datetime_series(s, sample=20):
    cnt = 0
    for v in s.dropna().astype(str).head(sample):
        vv = v.strip().replace("年","-").replace("月","-").replace("日","")
        m = re.search(r"20\d{2}[-/\\.]\d{1,2}[-/\\.]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?", vv)
        if m:
            cnt += 1
        else:
            try:
                _ = dateparser.parse(vv)
                cnt += 1
            except Exception:
                pass
    n = min(sample, s.dropna().shape[0])
    return cnt >= max(3, n // 2)

def _looks_like_amount_series(s, sample=20):
    cnt = 0
    for v in s.dropna().astype(str).head(sample):
        vv = v.strip().replace(",", "").replace("，", "")
        if re.fullmatch(r"[-+]?\(?\d+(?:\.\d+)?\)?", vv) or re.fullmatch(r"[-+]?\d{1,3}(?:[,，]\d{3})*(?:\.\d+)?", v.strip()):
            cnt += 1
    n = min(sample, s.dropna().shape[0])
    return cnt >= max(3, n // 2)

def _is_generic_table(df):
    cols_norm = _normalize_cols(df.columns)
    # 直接列名有命中
    has_date_name = any(any(h in c for h in DATE_HEADERS) for c in cols_norm)
    has_amt_name = any(any(h in c for h in AMOUNT_HEADERS) for c in cols_norm)
    has_bal_name = any(any(h in c for h in BALANCE_HEADERS) for c in cols_norm)
    has_desc_name = any(any(h in c for h in DESCRIPTION_HEADERS) for c in cols_norm)
    has_dir_name = any(any(h in c for h in DIRECTION_HEADERS) for c in cols_norm)
    if sum([has_date_name, has_amt_name or has_bal_name, has_desc_name or has_dir_name]) >= 2:
        return True
    # 无明显列名 → 以列内容猜测
    ok_date = False; ok_amt = False
    for c in df.columns:
        s = df[c]
        try:
            if not ok_date and _looks_like_datetime_series(s):
                ok_date = True
            if not ok_amt and _looks_like_amount_series(s):
                ok_amt = True
        except Exception:
            continue
    return ok_date and ok_amt

def _cell_is_number_like(s: str) -> bool:
    s = str(s).strip()
    if not s:
        return False
    s2 = s.replace(",", "").replace("，", "")
    return bool(re.fullmatch(r"[-+]?\(?\d+(?:\.\d+)?\)?", s2)) or bool(re.fullmatch(r"[-+]?\d{1,3}(?:[,，]\d{3})*(?:\.\d+)?", s))

def _row_score_as_header(cells, next_cells=None):
    """Heuristic score for choosing a header row among the first N rows.
    Favor rows with mostly text (not numeric), many non-empty unique labels.
    If next row exists and is more numeric-like, boost the score.
    """
    vals = [str(x) if x is not None else "" for x in cells]
    non_empty = [v for v in vals if v.strip() != ""]
    if not non_empty:
        return 0.0
    unique_ratio = len(set(non_empty)) / max(1, len(non_empty))
    numeric_ratio = sum(1 for v in non_empty if _cell_is_number_like(v)) / len(non_empty)
    # header-like tokens slight boost
    norm = _normalize_cols(vals)
    tokens = 0
    key_sets = [DATE_HEADERS, AMOUNT_HEADERS, BALANCE_HEADERS, DESCRIPTION_HEADERS, DIRECTION_HEADERS, METHOD_HEADERS, TXN_ID_HEADERS, MERCHANT_ID_HEADERS]
    for c in norm:
        if any(any(h in c for h in ks) for ks in key_sets):
            tokens += 1
    token_ratio = tokens / max(1, len(non_empty))
    header_density = len(non_empty) / max(1, len(vals))
    score = unique_ratio * 2.0 + (1.0 - numeric_ratio) * 1.0 + token_ratio * 0.5 + header_density * 0.6
    # 过于稀疏的一行（类似“人民币 总支出 …”整行合并）扣分
    if len(non_empty) <= max(2, len(vals)//6):
        score *= 0.4
    if next_cells is not None:
        nvals = [str(x) if x is not None else "" for x in next_cells]
        n_non_empty = [v for v in nvals if v.strip() != ""]
        if n_non_empty:
            n_numeric_ratio = sum(1 for v in n_non_empty if _cell_is_number_like(v)) / len(n_non_empty)
            # If next row is more numeric (data row), consider this current row a real header
            score += max(0.0, n_numeric_ratio - numeric_ratio) * 0.8
    return score

def _guess_header_row(raw_df, max_scan: int = 10):
    """Pick the most probable header row in the first few rows without assuming schemas.
    Returns index or None if no good candidate.
    """
    n = len(raw_df)
    if n == 0:
        return None
    limit = min(max_scan, n)
    best_idx, best_score = None, 0.0
    for ridx in range(limit):
        cells = list(raw_df.iloc[ridx])
        next_cells = list(raw_df.iloc[ridx + 1]) if (ridx + 1 < n) else None
        score = _row_score_as_header(cells, next_cells)
        if score > best_score:
            best_score, best_idx = score, ridx
    # Use a mild threshold to avoid picking empty/noisy top rows
    return best_idx if best_score >= 0.6 else None

def _combine_header_rows(row1, row2):
    """将上下两行表头合并为一行。尽量保持原词，短词做拼接，如 '借' + '贷' → '借/贷'。"""
    out = []
    n = max(len(row1), len(row2))
    for i in range(n):
        a = str(row1[i]).strip() if i < len(row1) and row1[i] is not None else ""
        b = str(row2[i]).strip() if i < len(row2) and row2[i] is not None else ""
        if a and b:
            if (a, b) in [("借", "贷"), ("贷", "借")]:
                out.append("借/贷")
            elif (a, b) in [("钞", "汇"), ("汇", "钞")]:
                out.append("钞/汇")
            elif len(a) <= 6 and len(b) <= 6:
                out.append(f"{a}/{b}")
            else:
                out.append(a if len(a) >= len(b) else b)
        else:
            out.append(a or b)
    return out

def _maybe_two_row_header(raw_df, header_idx):
    """若 header 下一行仍像表头（短词、少数字），尝试合并为双行表头。返回最终 header 列表。"""
    header_raw = _raw_header_cells(list(raw_df.iloc[header_idx]))
    if header_idx + 1 >= len(raw_df):
        return header_raw
    next_row = list(raw_df.iloc[header_idx + 1])
    # 如果下一行数字占比很低且含有典型短词（如 借/贷、钞/汇）则合并
    nvals = [str(x).strip() for x in next_row]
    n_non_empty = [v for v in nvals if v]
    if not n_non_empty:
        return header_raw
    n_numeric_ratio = sum(1 for v in n_non_empty if _cell_is_number_like(v)) / len(n_non_empty)
    short_token_cnt = sum(1 for v in n_non_empty if len(v) <= 4)
    contains_bank_tokens = any(any(tok in v for tok in BANK_GENERIC_HEADERS) for v in n_non_empty)
    # 双表头启发：短词占比较高、数字占比低，或含典型银行表头碎词
    if n_numeric_ratio <= 0.2 and (short_token_cnt >= max(2, len(n_non_empty)//3) or contains_bank_tokens):
        return _combine_header_rows(header_raw, _raw_header_cells(next_row))
    return header_raw

def _merge_wrapped_rows_by_serial(df: pd.DataFrame) -> pd.DataFrame:
    """按“序号”列把被换行拆成多行的记录合并。
    规则：若当前行序号为空，且上一行有有效序号，则认为是上一条记录续行；
    文本列进行拼接（以换行连接），数值/日期列若为空则向上补值。
    """
    if "序号" not in df.columns:
        return df
    ser = df["序号"].astype(str).str.strip()
    # 允许序号列带小数/非数字混杂，仅以是否为空判定续行
    out_rows = []
    buffer = None
    text_like_cols = [c for c in df.columns if df[c].dtype == object or c in ["摘要","交易地点/渠道","对手账号/户名","对手账号","对手户名"]]
    for _, row in df.iterrows():
        idxv = str(row.get("序号", "")).strip()
        if idxv:
            if buffer is not None:
                out_rows.append(buffer)
            buffer = row.copy()
        else:
            if buffer is None:
                buffer = row.copy()
            else:
                # 合并文本列
                for c in text_like_cols:
                    a = str(buffer.get(c, "") if buffer.get(c) is not None else "").strip()
                    b = str(row.get(c, "") if row.get(c) is not None else "").strip()
                    if b:
                        buffer[c] = (a + ("\n" if a else "") + b) if a else b
                # 数值/日期列补值
                for c in df.columns:
                    if c in text_like_cols:
                        continue
                    if pd.isna(buffer.get(c)) or str(buffer.get(c)).strip() == "":
                        v = row.get(c)
                        if v is not None and str(v).strip() != "":
                            buffer[c] = v
    if buffer is not None:
        out_rows.append(buffer)
    return pd.DataFrame(out_rows, columns=df.columns)

def _clean_money(x):
    if pd.isna(x): return None
    s = str(x).strip().replace(",", "").replace("，","")
    m = re.search(r"[-+]?[\d]+(?:\.\d+)?", s)
    return float(m.group()) if m else None

def _clean_money_signed(x):
    if pd.isna(x): return None
    s = str(x).strip().replace(",", "").replace("，", "")
    neg = False
    if "(" in s and ")" in s:
        neg = True
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    m = re.search(r"[-+]?[\d]+(?:\.\d+)?", s)
    if not m:
        return None
    v = float(m.group())
    if s.startswith("-"):
        neg = True
    return -abs(v) if neg else v

def _parse_dt(x):
    if pd.isna(x): return None
    s = str(x).strip().replace("年","-").replace("月","-").replace("日","")
    try:
        return pd.to_datetime(dateparser.parse(s))
    except Exception:
        return None

def _standardize(df):
    # 合并重复列工具：将同名列按“前者优先、缺失用后者填充”合并为单列
    def _merge_duplicate_columns(d):
        from collections import defaultdict
        cols = list(d.columns)
        pos = defaultdict(list)
        for i, c in enumerate(cols):
            pos[c].append(i)
        for c, idxs in pos.items():
            if len(idxs) > 1:
                base = d.iloc[:, idxs[0]]
                for j in idxs[1:]:
                    base = base.combine_first(d.iloc[:, j])
                d.iloc[:, idxs[0]] = base
        # 去重仅保留首次出现的列
        d = d.loc[:, ~pd.Index(d.columns).duplicated(keep='first')]
        return d

    rename_map = {}
    for c in df.columns:
        k = str(c).replace("\n","").replace(" ","")
        # 交易单号/交易号/流水号
        if any(h in k for h in TXN_ID_HEADERS):
            rename_map[c] = "交易单号"
        elif any(h in k for h in DATE_HEADERS): rename_map[c] = "交易时间"
        elif "交易类型" in k or "业务类型" in k: rename_map[c] = "交易类型"
        elif any(h in k for h in DIRECTION_HEADERS): rename_map[c] = "收/支/其他"
        elif any(h in k for h in METHOD_HEADERS): rename_map[c] = "交易方式"
        elif any(h in k for h in ["借方金额","借方发生额","借(入)"]): rename_map[c] = "借方金额"
        elif any(h in k for h in ["贷方金额","贷方发生额","贷(出)"]): rename_map[c] = "贷方金额"
        elif any(h in k for h in ["收入金额"]): rename_map[c] = "收入金额"
        elif any(h in k for h in ["支出金额"]): rename_map[c] = "支出金额"
        elif any(h in k for h in AMOUNT_HEADERS): rename_map[c] = "金额(元)"
        elif any(h in k for h in COUNTERPARTY_HEADERS):
            rename_map[c] = "交易对方"
        elif any(h in k for h in MERCHANT_ID_HEADERS): rename_map[c] = "商户单号"
        elif any(h in k for h in DESCRIPTION_HEADERS): rename_map[c] = "备注"
        elif any(h in k for h in BALANCE_HEADERS): rename_map[c] = "余额"
        elif "资金渠道" in k or "交易渠道" in k or "渠道" in k: rename_map[c] = "资金渠道"
        elif "手续费" in k: rename_map[c] = "手续费"
        else:
            rename_map[c] = c
    df = df.rename(columns=rename_map)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df = _merge_duplicate_columns(df)

    # 金额正负与借/贷专列处理
    has_debit = "借方金额" in df.columns
    has_credit = "贷方金额" in df.columns
    has_income = "收入金额" in df.columns
    has_expend = "支出金额" in df.columns

    if has_debit:
        df["借方金额"] = df["借方金额"].apply(_clean_money)
    if has_credit:
        df["贷方金额"] = df["贷方金额"].apply(_clean_money)
    if has_income:
        df["收入金额"] = df["收入金额"].apply(_clean_money)
    if has_expend:
        df["支出金额"] = df["支出金额"].apply(_clean_money)

    if "交易时间" in df.columns:
        df["交易时间"] = df["交易时间"].apply(_parse_dt)

    if "金额(元)" in df.columns:
        # 若存在括号负号等，提取带符号金额
        df["金额(元)"] = df["金额(元)"].apply(_clean_money)

    if "收/支/其他" in df.columns and "金额(元)" in df.columns:
        def signed(row):
            v = row.get("金额(元)")
            t = str(row.get("收/支/其他","")).strip()
            if v is None: return None
            if "支出" in t: return -abs(v)
            if "收入" in t: return abs(v)
            return v
        df["金额(带符号)"] = df.apply(signed, axis=1)
    # 借/贷 或 收入/支出列 → 直接生成带符号金额
    if "金额(带符号)" not in df.columns:
        if has_debit or has_credit:
            debit = df.get("借方金额")
            credit = df.get("贷方金额")
            if debit is not None and credit is not None:
                df["金额(带符号)"] = (credit.fillna(0) - debit.fillna(0)).astype(float)
                df["金额(元)"] = df["金额(带符号)"].abs()
                df["收/支/其他"] = df["金额(带符号)"].apply(lambda v: "收入" if v is not None and v>=0 else "支出")
            elif debit is not None:
                df["金额(带符号)"] = -debit
                df["金额(元)"] = debit.abs()
                df["收/支/其他"] = "支出"
            elif credit is not None:
                df["金额(带符号)"] = credit
                df["金额(元)"] = credit.abs()
                df["收/支/其他"] = "收入"
        elif has_income or has_expend:
            income = df.get("收入金额")
            expend = df.get("支出金额")
            income = income if income is not None else 0
            expend = expend if expend is not None else 0
            df["金额(带符号)"] = (pd.Series(income).fillna(0) - pd.Series(expend).fillna(0)).astype(float)
            df["金额(元)"] = df["金额(带符号)"].abs()
            df["收/支/其他"] = df["金额(带符号)"].apply(lambda v: "收入" if v is not None and v>=0 else "支出")

    mapping = {
        "交易单号":"transaction_id",
        "交易时间":"timestamp",
        "交易类型":"type",
        "收/支/其他":"direction",
        "交易方式":"method",
        "金额(元)":"amount",
        "金额(带符号)":"amount_signed",
        "交易对方":"counterparty",
        "商户单号":"merchant_order_id",
        "备注":"note",
        "余额":"balance",
        "资金渠道":"channel",
        "手续费":"fee",
    }
    for zh, en in mapping.items():
        if zh in df.columns and en not in df.columns:
            df[en] = df[zh]

    def not_data(r):
        row_str = " ".join(str(x) for x in r.values)
        return bool(re.search(r"(合计|余额|人民币/单位|仅作|证明|章|交易明细对应时间段)", row_str))
    df = df[~df.apply(not_data, axis=1)]

    if "timestamp" in df.columns:
        df = df.sort_values("timestamp").reset_index(drop=True)
    return df

def _merge_duplicate_columns_public(d: pd.DataFrame) -> pd.DataFrame:
    from collections import defaultdict
    cols = list(d.columns)
    pos = defaultdict(list)
    for i, c in enumerate(cols):
        pos[c].append(i)
    for c, idxs in pos.items():
        if len(idxs) > 1:
            base = d.iloc[:, idxs[0]]
            for j in idxs[1:]:
                base = base.combine_first(d.iloc[:, j])
            d.iloc[:, idxs[0]] = base
    d = d.loc[:, ~pd.Index(d.columns).duplicated(keep='first')]
    return d

def _filter_not_data_rows(df: pd.DataFrame) -> pd.DataFrame:
    def not_data(r):
        row_str = " ".join(str(x) for x in r.values)
        return bool(re.search(r"(合计|余额|人民币/单位|仅作|证明|章|交易明细对应时间段)", row_str))
    if df.empty:
        return df
    return df[~df.apply(not_data, axis=1)]

# ------------------ 表格读取（文本/可检索PDF） ------------------

def read_by_camelot(pdf_path):
    if camelot is None:
        return []
    try:
        dfs = []
        # 优先 lattice（网格线表格更稳），扩大 line_scale 以合并碎线
        try:
            tables = camelot.read_pdf(
                pdf_path, pages="all",
                flavor="lattice", strip_text=" \n",
                edge_tol=200, row_tol=8, column_tol=8, line_scale=40
            )
            dfs.extend([t.df for t in tables] if tables else [])
        except Exception:
            pass
        # 再尝试 stream 兜底
        if not dfs:
            try:
                tables = camelot.read_pdf(
                    pdf_path, pages="all",
                    flavor="stream", strip_text=" \n",
                    row_tol=12, column_tol=8
                )
                dfs.extend([t.df for t in tables] if tables else [])
            except Exception:
                pass
        fixed = []
        for raw in dfs:
            if raw.shape[0] >= 2:
                # 搜索最可能的表头（泛化 + 兜底）
                header_row_idx = None
                limit = min(8, len(raw))
                for ridx in range(limit):
                    header_norm = _normalize_cols(list(raw.iloc[ridx]))
                    if (sum(1 for h in DATE_HEADERS if any(h in c for c in header_norm)) >= 1 and 
                        (sum(1 for h in AMOUNT_HEADERS if any(h in c for c in header_norm)) >= 1 or any("借" in c or "贷" in c for c in header_norm))):
                        header_row_idx = ridx
                        break
                if header_row_idx is None:
                    # 通用启发式猜测
                    header_row_idx = _guess_header_row(raw)
                if header_row_idx is None:
                    header_row_idx = 0
                header_raw = _maybe_two_row_header(raw, header_row_idx)
                body = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
                if body.shape[1] == len(header_raw):
                    body.columns = header_raw
                    body = _merge_wrapped_rows_by_serial(body)
                    fixed.append(body)
        return fixed
    except Exception:
        return []

def read_by_tabula(pdf_path):
    if tabula is None:
        return []
    try:
        dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, lattice=True, guess=True)
        if not dfs:
            dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, stream=True, guess=True)
        return dfs or []
    except Exception:
        return []

def read_by_pdfplumber(pdf_path):
    out = []
    last_header_raw = None
    last_header_norm = None
    strategies = [
        {},
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "edge_min_length": 40,
            "min_words_vertical": 1,
            "min_words_horizontal": 1,
            "keep_blank_chars": True,
            "intersection_x_tolerance": 5,
            "intersection_y_tolerance": 5,
        },
        {
            "vertical_strategy": "lines_strict",
            "horizontal_strategy": "lines_strict",
            "snap_tolerance": 2,
            "join_tolerance": 2,
            "edge_min_length": 30,
            "keep_blank_chars": True,
        },
        {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "keep_blank_chars": True,
        },
    ]
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages:
                page_tables = []
                for ts in strategies:
                    try:
                        tbls = p.extract_tables(table_settings=ts) or []
                    except Exception:
                        tbls = []
                    if not tbls:
                        continue
                    for t in tbls:
                        df = pd.DataFrame(t)
                        if not df.empty:
                            page_tables.append(df)
                    # 若已获得较大表，停止尝试更多策略以节省时间
                    if any(d.shape[0] >= 10 and d.shape[1] >= 5 for d in page_tables):
                        break

                for df in page_tables:
                    # 在前几行中查找真正的表头（先按已知线索，再按启发式）
                    header_row_idx = None
                    limit = min(8, len(df))
                    for ridx in range(limit):
                        header_norm = _normalize_cols(list(df.iloc[ridx]))
                        # 同时兼容微信/支付宝/通用流水表头
                        if (sum(1 for h in WECHAT_HEADERS if h in header_norm) >= 5) or \
                           (sum(1 for h in ALIPAY_HEADERS if h in header_norm) >= 5) or \
                           (sum(1 for h in DATE_HEADERS if any(h in c for c in header_norm)) >= 1 and \
                            (sum(1 for h in AMOUNT_HEADERS if any(h in c for c in header_norm)) >= 1 or \
                             any("借/贷" in c or "摘要" in c or "余额" in c for c in header_norm))):
                            header_row_idx = ridx
                            break
                        # 少量列名带“日期/摘要/借/贷/余额”的组合也视为表头
                        if any("日期" in c or "时间" in c for c in header_norm) and any("摘要" in c for c in header_norm):
                            header_row_idx = ridx
                            break
                    # 启发式猜测（任意表头）
                    if header_row_idx is None:
                        header_row_idx = _guess_header_row(df)
                    # 如果只有 1-2 行且猜到第 0 行为“表头”，但上一页已有表头且该行更像数据，则视为数据行
                    if header_row_idx == 0 and len(df) <= 3 and last_header_raw is not None:
                        vals = [str(x).strip() for x in list(df.iloc[0])]
                        non_empty = [v for v in vals if v]
                        num_ratio = sum(1 for v in non_empty if _cell_is_number_like(v)) / max(1, len(non_empty))
                        has_date_like = any(re.search(r"20\d{2}[-/\\.]?\d{1,2}[-/\\.]?\d{1,2}", v) for v in non_empty)
                        if num_ratio >= 0.3 or has_date_like:
                            header_row_idx = None
                    if header_row_idx is not None:
                        header_raw = _maybe_two_row_header(df, header_row_idx)
                        header_norm = _normalize_cols(list(header_raw))
                        body = df.iloc[header_row_idx + 1 :].reset_index(drop=True)
                        if len(header_raw) == body.shape[1]:
                            body.columns = header_raw  # 导出保持原表头
                            body = _merge_wrapped_rows_by_serial(body)
                            out.append(body)
                            last_header_raw = header_raw
                            last_header_norm = header_norm
                            continue
                    # 若本页未找到表头，但此前已识别过表头，且列数匹配，则沿用上一页表头（保持原表头）
                    if last_header_raw is not None and len(last_header_raw) == df.shape[1]:
                        body = df.reset_index(drop=True)
                        body.columns = last_header_raw
                        out.append(body)
                        continue
                    # 最后兜底：将第 0 行视作表头
                    if len(df) >= 1 and last_header_raw is None:
                        header_raw = _raw_header_cells(list(df.iloc[0]))
                        body = df.iloc[1:].reset_index(drop=True)
                        if body.shape[1] == len(header_raw):
                            body.columns = header_raw
                            out.append(body)
    except Exception:
        pass
    return out

def read_by_pdfplumber_words(pdf_path):
    """使用基于文字坐标的列切分重建表格，适用于 pdfplumber 表格识别只获得表头的情况。"""
    outs = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages:
                words = p.extract_words(keep_blank_chars=False)
                if not words:
                    continue
                # 1) 按 y 聚类为行
                lines = []
                words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))
                cur = []
                last_top = None
                for w in words_sorted:
                    top = w["top"]
                    if last_top is None or abs(top - last_top) <= 3:
                        cur.append(w)
                        last_top = top if last_top is None else (last_top*0.7 + top*0.3)
                    else:
                        if cur:
                            lines.append(cur)
                        cur = [w]
                        last_top = top
                if cur:
                    lines.append(cur)

                # 2) 找包含表头关键字的行
                header_idx = None
                header_words = None
                key_tokens = ["序号","摘要","交易日期","交易金额","账户余额","对方账号","户名","对方账号与户名","币别","钞汇","交易地点","附言"]
                for i, line in enumerate(lines[:30]):
                    txts = [t["text"].strip() for t in line]
                    score = sum(1 for k in key_tokens if any(k in t for t in txts))
                    if any("序号" in t for t in txts) and score >= 3:
                        header_idx = i
                        header_words = line
                        break
                if header_idx is None:
                    continue
                # 3) 由表头 token 的 x 中心确定列边界
                headers = []
                centers = []
                for t in header_words:
                    tx = t["text"].strip()
                    if not tx:
                        continue
                    headers.append(tx)
                    centers.append((t["x0"] + t["x1"]) / 2.0)
                if len(headers) < 3:
                    continue
                order = sorted(range(len(centers)), key=lambda i: centers[i])
                centers = [centers[i] for i in order]
                headers = [headers[i] for i in order]
                bounds = []
                # 左右边界：用页面的 2%/98% 宽度近似
                x_min, x_max = p.cropbox[0], p.cropbox[2]
                if x_min is None: x_min = 0
                if x_max is None: x_max = max(centers) + 5
                bounds.append(x_min + 2)
                for i in range(len(centers)-1):
                    bounds.append((centers[i] + centers[i+1]) / 2.0)
                bounds.append(x_max - 2)

                # 4) 收集数据行（表头行以下）
                data_lines = lines[header_idx+1:]
                rows = []
                for line in data_lines:
                    # 过滤页脚/统计
                    ltxt = " ".join(t["text"] for t in line)
                    if re.search(r"(合计|小计|人民币/单位|仅作|证明|章)", ltxt):
                        continue
                    cols = ["" for _ in headers]
                    for w in line:
                        cx = (w["x0"] + w["x1"]) / 2.0
                        # 找到所在列
                        j = 0
                        while j < len(bounds)-1 and not(bounds[j] <= cx < bounds[j+1]):
                            j += 1
                        if j >= len(headers):
                            j = len(headers)-1
                        s = w["text"].strip()
                        if s:
                            cols[j] = (cols[j] + (" " if cols[j] else "") + s)
                    # 粗判数据行（序号列或日期列有值）
                    if any(cols) and (re.match(r"\d+$", cols[0]) or re.search(r"20\d{2}[-/]?\d{1,2}[-/]?\d{1,2}", " ".join(cols))):
                        rows.append(cols)
                if not rows:
                    continue
                df = pd.DataFrame(rows, columns=headers)
                # 修补：将“序号”中误入的摘要词移到“摘要”列
                try:
                    if "序号" in df.columns and any(c in df.columns for c in ["摘要","摘要/用途"]):
                        sum_col = "摘要" if "摘要" in df.columns else "摘要/用途"
                        def split_seq(x):
                            s = str(x).strip()
                            m = re.match(r"^(\d+)\s+(.+)$", s)
                            return (m.group(1), m.group(2)) if m else (s, None)
                        seqs = df["序号"].apply(split_seq)
                        df["序号"] = seqs.apply(lambda t: t[0])
                        df[sum_col] = [b if (str(df.at[i, sum_col]).strip()=="" and b) else df.at[i, sum_col] for i, (_, b) in enumerate(seqs)]
                except Exception:
                    pass
                outs.append(df)
    except Exception:
        return []
    return outs

# ------------------ OCR 路线 A：OCRmyPDF ------------------

def has_ocrmypdf():
    try:
        subprocess.run(["ocrmypdf", "--version"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True
    except Exception:
        return False

def ocrmypdf_to_searchable(src_pdf):
    """用 OCRmyPDF 转换为可检索 PDF（保留版面），返回新PDF路径"""
    if not has_ocrmypdf():
        return None
    tmp_out = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False).name
    # --skip-text: 已有文字的不重复OCR；--language chi_sim+eng；--output-type pdf
    cmd = ["ocrmypdf", "--skip-text", "-l", "chi_sim+eng", src_pdf, tmp_out]
    res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if res.returncode != 0:
        return None
    return tmp_out

def preclean_red_stamp_then_ocr(src_pdf):
    """去红章 + 二值化 → 合成图像PDF → OCR 生成可检索PDF。
    适用于红章覆盖文字导致文本缺失的情况。
    返回新PDF路径；若依赖缺失则返回 None。
    """
    if img2pdf is None or pytesseract is None:
        # 没有图像或OCR依赖，直接退回 ocrmypdf 原始处理
        return ocrmypdf_to_searchable(src_pdf)
    try:
        pages = convert_from_path(src_pdf, dpi=300)
        if not pages:
            return None
        cleaned_imgs = []
        for p in pages:
            bgr = cv2.cvtColor(np.array(p), cv2.COLOR_RGB2BGR)
            bin_img = remove_red_stamp_and_binarize(bgr)
            cleaned_imgs.append(cv2.cvtColor(bin_img, cv2.COLOR_GRAY2RGB))
        # 写到临时 PDF（仅图像）
        tdir = tempfile.mkdtemp()
        img_paths = []
        for i, img in enumerate(cleaned_imgs):
            ip = os.path.join(tdir, f"p{i:03d}.jpg")
            cv2.imwrite(ip, img, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
            img_paths.append(ip)
        tmp_img_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False).name
        with open(tmp_img_pdf, "wb") as f:
            f.write(img2pdf.convert(img_paths))
        # OCR 成可检索 PDF
        out_pdf = ocrmypdf_to_searchable(tmp_img_pdf)
        try:
            os.remove(tmp_img_pdf)
        except Exception:
            pass
        return out_pdf
    except Exception:
        return None

# ------------------ OCR 路线 B：逐页图片 + 红章处理 ------------------

def remove_red_stamp_and_binarize(img_bgr):
    """
    1) HSV分割红色（两段Hue），得到mask
    2) 形态学膨胀填补孔洞
    3) inpaint（Telea）修复红章区域
    4) 转灰度 + 自适应阈值 + 轻度去噪
    """
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    lower_red1 = np.array([0, 70, 60]); upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 70, 60]); upper_red2 = np.array([180, 255, 255])
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask = cv2.bitwise_or(mask1, mask2)
    kernel = np.ones((5,5), np.uint8)
    mask = cv2.dilate(mask, kernel, iterations=1)
    inpainted = cv2.inpaint(img_bgr, mask, 3, cv2.INPAINT_TELEA)
    gray = cv2.cvtColor(inpainted, cv2.COLOR_BGR2GRAY)
    # 轻度去噪 + 自适应阈值
    gray = cv2.fastNlMeansDenoising(gray, None, 15, 7, 21)
    th = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                               cv2.THRESH_BINARY, 35, 10)
    return th

def deskew(binary_img):
    # 简单霍夫/轮廓角度估计，做轻微矫正
    coords = np.column_stack(np.where(binary_img < 255))
    if coords.size == 0:
        return binary_img
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
    (h, w) = binary_img.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    rotated = cv2.warpAffine(binary_img, M, (w, h), flags=cv2.INTER_LINEAR,
                             borderMode=cv2.BORDER_REPLICATE)
    return rotated

def ocr_pages_to_rows(pdf_path):
    """逐页OCR并按行粗解析成交易记录（兜底方案）"""
    if pytesseract is None:
        return pd.DataFrame()
    pages = convert_from_path(pdf_path, dpi=350)  # 足够清晰
    rows = []
    dt_pat = r"(20\d{2}[-/\.年]\d{1,2}[-/\.月]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)"
    amt_pat = r"([-+]?\d{1,3}(?:[,，]\d{3})*(?:\.\d{1,2})?)"
    for pil_img in pages:
        img_bgr = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
        bin_img = remove_red_stamp_and_binarize(img_bgr)
        bin_img = deskew(bin_img)
        # 直接OCR
        txt = pytesseract.image_to_string(bin_img, lang="chi_sim+eng")
        # 逐行启发式抽取（根据你样张：每行包含 时间 与 金额）
        for line in txt.splitlines():
            line = line.strip()
            if not line: continue
            m_dt = re.search(dt_pat, line)
            m_amt = re.search(amt_pat, line)
            if m_dt and m_amt:
                rows.append({
                    "交易时间": m_dt.group(1),
                    "金额(元)": m_amt.group(1),
                    "收/支/其他": "未知",
                    "行文本": line
                })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # 尝试从行文本推断“收入/支出/方式/对方”（可按你实际样本再加规则）
    def infer_dir(s):
        s = str(s)
        if "支出" in s: return "支出"
        if "收入" in s: return "收入"
        return "其他"
    df["收/支/其他"] = df["行文本"].apply(infer_dir)
    return df.drop(columns=["行文本"])
# ------------------ 主流程 ------------------

def try_tables(pdf_path):
    def accept(df):
        return _is_wechat_table(df) or _is_alipay_table(df) or _is_generic_table(df)

    def rows_total(dfs):
        return sum(max(0, len(d)) for d in dfs)

    def cols_stats(dfs):
        cols = [d.shape[1] for d in dfs if getattr(d, 'shape', (0,0))[1] > 0]
        return (min(cols) if cols else 0, max(cols) if cols else 0, (sorted(cols)[len(cols)//2] if cols else 0))

    def score(dfs):
        if not dfs:
            return 0
        # 质量：更多列、更多行、更像通用流水
        r = rows_total(dfs)
        mn, mx, md = cols_stats(dfs)
        like_cnt = 0
        for d in dfs:
            cols = _normalize_cols(d.columns)
            bank_hits = sum(1 for k in BANK_GENERIC_HEADERS if any(k in c for c in cols))
            if _is_generic_table(d) or _is_wechat_table(d) or _is_alipay_table(d) or bank_hits >= 3:
                like_cnt += 1
        s = r + md * 360 + like_cnt * 240
        if md <= 2:
            s -= 600
        return s

    # 取各引擎的候选
    # 银行对账单常见为网格线表格，优先尝试 camelot
    cm = [df for df in read_by_camelot(pdf_path) if accept(df)]
    pp = [df for df in read_by_pdfplumber(pdf_path) if accept(df)]
    tb = [df for df in read_by_tabula(pdf_path) if accept(df)]

    # 如果 pdfplumber 结果存在但全是空数据行，尝试基于文字坐标的还原
    if pp and all(getattr(d, 'shape', (0,0))[0] == 0 for d in pp):
        pp_words = [df for df in read_by_pdfplumber_words(pdf_path) if accept(df)]
        if pp_words:
            pp = pp_words

    # 评估分数，选最佳引擎
    groups = {"camelot": cm, "pdfplumber": pp, "tabula": tb}
    best_name = max(groups, key=lambda k: score(groups[k]))
    best = groups[best_name]

    # 若最佳仍过于粗糙（列数普遍很少），尝试合并另一个更优列数的结果
    _, best_mx, best_md = cols_stats(best)
    if best_md <= 2:
        others = [(k, v) for k, v in groups.items() if k != best_name]
        others.sort(key=lambda kv: cols_stats(kv[1])[2], reverse=True)
        for _, v in others:
            if cols_stats(v)[2] > best_md:
                best = v
                break

    # 过滤异常：当存在 >=4 列的表时，丢弃只有 1-2 列的表
    _, mx, md = cols_stats(best)
    if md >= 4:
        best = [d for d in best if d.shape[1] >= 4]

    return best

def wechat_pdf_to_excel(pdf_path, out_xlsx="wechat_payments.xlsx", preserve_original_headers=True):
    # 1) 直接尝试表格抓取
    candidates = try_tables(pdf_path)

    # 2) 不行 → OCRmyPDF 生成可检索PDF，再抓表
    if not candidates:
        ocr_pdf = ocrmypdf_to_searchable(pdf_path)
        if ocr_pdf:
            candidates = try_tables(ocr_pdf)
            try:
                os.remove(ocr_pdf)
            except Exception:
                pass

    # 3) 可疑（大量文本列空白）→ 去红章 + OCR 后再抓表
    def _blank_ratio_on_cols(d: pd.DataFrame, keys):
        if d is None or d.empty:
            return 1.0
        cols = [c for c in d.columns if any(k in str(c) for k in keys)]
        if not cols:
            return 0.0
        sub = d[cols]
        tot = sub.shape[0] * sub.shape[1]
        if tot == 0:
            return 1.0
        empties = 0
        for c in cols:
            s = sub[c].astype(str)
            mask = (s == "") | (s.str.strip() == "") | (s.str.lower() == "nan")
            empties += mask.sum()
        return empties / tot

    suspect = False
    if candidates:
        ratios = []
        for d in candidates:
            ratios.append(_blank_ratio_on_cols(d, ["摘要","对方","户名","附言","用途","交易地点"]))
        suspect = any(r > 0.5 for r in ratios)

    if suspect:
        cleaned_pdf = preclean_red_stamp_then_ocr(pdf_path)
        if cleaned_pdf:
            better = try_tables(cleaned_pdf)
            # 若改进明显（文本空白率下降），采用改进后的结果
            def best_ratio(ds):
                return min((_blank_ratio_on_cols(x, ["摘要","对方","户名","附言","用途","交易地点"]) for x in ds), default=1)
            if best_ratio(better) + 1e-6 < best_ratio(candidates):
                candidates = better
            try:
                os.remove(cleaned_pdf)
            except Exception:
                pass

    # 4) 还不行 → 逐页OCR（红章处理）
    if not candidates:
        ocr_df = ocr_pages_to_rows(pdf_path)
        if not ocr_df.empty:
            candidates = [ocr_df]

    if not candidates:
        raise RuntimeError("未识别到交易数据。请确认PDF清晰度或提供样例让我细化规则。")

    std = [_standardize(df.copy()) for df in candidates]
    merged = pd.concat(std, ignore_index=True)

    # 原表头导出集（不做字段改名，仅做基础清理）
    orig_frames = []
    for df in candidates:
        d = df.copy()
        d = d.dropna(axis=1, how="all").dropna(axis=0, how="all")
        d = _merge_duplicate_columns_public(d)
        d = _filter_not_data_rows(d)
        orig_frames.append(d)
    orig_merged = pd.concat(orig_frames, ignore_index=True) if orig_frames else pd.DataFrame()
    # 尽量去重（按原始列全行对比）
    if not orig_merged.empty:
        orig_merged = orig_merged.drop_duplicates()

    # 去重
    dup_cols = [c for c in ["transaction_id","timestamp","amount_signed"] if c in merged.columns]
    if dup_cols:
        merged["__dup_key__"] = merged[dup_cols].astype(str).agg("|".join, axis=1)
        merged = merged.drop_duplicates("__dup_key__").drop(columns="__dup_key__")
    else:
        # 退化去重：尽量通过 (timestamp,amount,counterparty,note)
        key_cols = [c for c in ["timestamp","amount","counterparty","note"] if c in merged.columns]
        if key_cols:
            merged["__dup_key__"] = merged[key_cols].astype(str).agg("|".join, axis=1)
            merged = merged.drop_duplicates("__dup_key__").drop(columns="__dup_key__")

    # 导出（通用化 sheet 名称）
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # 明细表：保留原表字段名
        export_df = (orig_merged if preserve_original_headers else merged)
        export_df.to_excel(writer, sheet_name="Transactions", index=False)
        # 汇总表仍基于带符号金额进行统计（若可用）
        if "amount_signed" in merged.columns:
            tmp = merged.assign(direction=lambda d: d["amount_signed"].apply(lambda v: "收入" if v is not None and v>=0 else "支出"))
            summary = tmp.groupby("direction", dropna=False)["amount_signed"].sum().reset_index()
            summary.rename(columns={"amount_signed":"合计(元)"}, inplace=True)
            summary.to_excel(writer, sheet_name="Summary", index=False)

        # 轻度排版：冻结首行、自动换行、为金额/余额列设定数字格式
        try:
            wb = writer.book
            ws = wb["Transactions"]
            ws.freeze_panes = "A2"
            # 自动调整宽度（估算）并设置换行
            for col_idx, col in enumerate(export_df.columns, start=1):
                max_len = max([len(str(col))] + [len(str(v)) for v in export_df[col].astype(str).head(200)])
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(60, max(10, int(max_len * 1.2)))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrap_text=True)
            # 数字格式
            money_like = [c for c in export_df.columns if any(k in str(c) for k in ["金额","余额"]) and not str(c).startswith("金额(带符号)")]
            from openpyxl.styles import numbers
            for c in money_like:
                j = export_df.columns.get_loc(c) + 1
                for i in range(2, ws.max_row + 1):
                    cell = ws.cell(row=i, column=j)
                    # 仅在数值型时设置格式
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        except Exception:
            pass

    return out_xlsx

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="交易流水PDF转Excel（保留原表字段名，含OCR兜底）")
    parser.add_argument("pdf", help="输入PDF路径")
    parser.add_argument("-o", "--out", default="wechat_payments.xlsx", help="输出Excel路径")
    parser.add_argument("--raw-headers", dest="raw_headers", action="store_true", help="保留原表字段名导出（默认）")
    parser.add_argument("--std-headers", dest="raw_headers", action="store_false", help="导出标准化字段名（不建议）")
    parser.set_defaults(raw_headers=True)
    args = parser.parse_args()
    path = wechat_pdf_to_excel(args.pdf, args.out, preserve_original_headers=args.raw_headers)
    print("已导出：", path)
